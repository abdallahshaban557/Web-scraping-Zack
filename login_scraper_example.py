#CTRL+ALT+N to Run the python script
from bs4 import BeautifulSoup
from datetime import datetime, timedelta
import requests
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from openpyxl import Workbook
from openpyxl import load_workbook

#x = urllib.request.urlopen('https://pythonprogramming.net')

try:

    #Stop remember me from showing up
    chrome_options = Options()
    chrome_options.add_experimental_option('prefs', {
        'credentials_enable_service': False,
        'profile': {
            'password_manager_enabled': False
        }
    })
    #Open the chrome driver, and navigate to the page
    driver = webdriver.Chrome("chromedriver/chromedriver.exe",chrome_options=chrome_options)
    driver.get("https://www.zacks.com/")

    

    #Load original Excel file - and select the needed sheet - get maximum row number
    #Result_File = load_workbook('results_file/messedupb.xlsx')
    Result_File = load_workbook('results_file/messedupb.xlsx')
    Result_sheet = Result_File.worksheets[0]
    Max_Row_Count = Result_sheet.max_row

    #Count not found ticker EPS on site
    Not_Found_Tickers = 0
    
    for i in range(2, Max_Row_Count):
        #Get the ticker and EPS date from data sheet
        print(i)  

        #Save a final results sheet
        Result_File.save('results_file/FINAL.xlsx')  

        Ticker_Symbol = Result_sheet.cell(row=i, column=6).value
        
        #Get date from the initial file. Minus and add a day for better results
        New_EPS_date_format = Result_sheet.cell(row=i, column=3).value.date()
        EPS_Day = New_EPS_date_format.strftime('%#d')
        EPS_Month = New_EPS_date_format.strftime('%#m')
        EPS_Year = New_EPS_date_format.strftime('%Y')
        EPS_Date = EPS_Month + "/" + EPS_Day + "/" + EPS_Year
        
        EPS_Date_plus_day = New_EPS_date_format + timedelta(days=1)
        EPS_Day = EPS_Date_plus_day.strftime('%#d')
        EPS_Month = EPS_Date_plus_day.strftime('%#m')
        EPS_Year = EPS_Date_plus_day.strftime('%Y')
        EPS_Date_plus_day = EPS_Month + "/" + EPS_Day + "/" + EPS_Year

        EPS_Date_minus_day = New_EPS_date_format - timedelta(days=1)
        EPS_Day = EPS_Date_minus_day.strftime('%#d')
        EPS_Month = EPS_Date_minus_day.strftime('%#m')
        EPS_Year = EPS_Date_minus_day.strftime('%Y')
        EPS_Date_minus_day = EPS_Month + "/" + EPS_Day + "/" + EPS_Year

        #Get all EPS values in data sheet
        IBES_Estimize_actual = Result_sheet.cell(row=i, column=2).value
        IBES_actual = Result_sheet.cell(row=i, column=4).value
        IBES_adj_actual = Result_sheet.cell(row=i, column=5).value

        #Locate the search box on top of the page
        Ticker_Symbol_Search = driver.find_element_by_name('search-q')

        #send ticker symbol value to the search box
        Ticker_Symbol_Search.send_keys(Ticker_Symbol)
        
        #Press enter on the search box
        Ticker_Symbol_Search.send_keys(Keys.ENTER)

        try:
            #find earnings announcement from the sidemenu
            Earnings_Announcement = driver.find_element_by_xpath('//*[@id="left_rail"]/nav/div[2]/ul[3]/li[2]/ul/li[4]/a')

            #Click on earnings announcement in the sidemenu
            driver.execute_script("arguments[0].click();", Earnings_Announcement)

            #Select the drop down row list and select 100 as the row number
            Drop_Down_List = Select(driver.find_element_by_xpath('//*[@id="earnings_announcements_earnings_table_length"]/label/select'))
            Drop_Down_List.select_by_visible_text('100')

            time.sleep(1)

            #Row number in the website EPS table
            EPS_Table_Row_Number = 1
        
            #Go through the whole table of EPS
            while True:
                try:
                    #Grab date from the website
                    Date_Of_EPS = driver.find_element_by_xpath('//*[@id="earnings_announcements_earnings_table_wrapper"]/div[3]/div[3]/div[2]/div/table/tbody/tr['+ str(EPS_Table_Row_Number) +']/td').text

                    #compare date from the data sheet with the website, and assign the right values based on EPS
                    if Date_Of_EPS == EPS_Date or Date_Of_EPS == EPS_Date_minus_day or Date_Of_EPS == EPS_Date_plus_day:
                        Reported_EPS = driver.find_element_by_xpath('//*[@id="earnings_announcements_earnings_table"]/tbody/tr[' + str(EPS_Table_Row_Number) + ']/td[4]').text
                        if Reported_EPS != "--":
                            Reported_EPS = Reported_EPS.replace("$","")
                            Reported_EPS = float(Reported_EPS)
                            Result_sheet.cell(row=i, column=11).value = Reported_EPS
                        else:
                            Result_sheet.cell(row=i, column=11).value = Reported_EPS

                        #Condition to add the value of IBES_RIGHT in data sheet
                        if Reported_EPS == "--":
                            Result_sheet.cell(row=i, column=9).value = 'N/A'
                        elif IBES_actual == Reported_EPS or IBES_adj_actual == Reported_EPS:
                            Result_sheet.cell(row=i, column=9).value = 1
                        elif IBES_Estimize_actual != Reported_EPS:
                            Result_sheet.cell(row=i, column=9).value = 0
                        else:
                            Result_sheet.cell(row=i, column=9).value = 0
                         
                        #Condition to add the value of ESTIMIZE to data sheet
                        if Reported_EPS == "--":
                            Result_sheet.cell(row=i, column=10).value = 'N/A'
                        elif IBES_Estimize_actual == Reported_EPS:
                            Result_sheet.cell(row=i, column=10).value = 1
                        else:
                            Result_sheet.cell(row=i, column=10).value = 0

                    #Increment counter for loop
                    EPS_Table_Row_Number += 1
                   
                
                except Exception as e:
                    print(str(e))
                    break         
        except Exception as e:
            Result_sheet.cell(row=i, column=9).value = 'N/A'
            Result_sheet.cell(row=i, column=10).value = 'N/A'
            Not_Found_Tickers += 1
            print('not found '+ str(Not_Found_Tickers))
            
    
    #Save a final results sheet
    Result_File.save('results_file/FINAL.xlsx')   

    #close the chrome page
    driver.close()

except Exception as e:
    print(str(e))